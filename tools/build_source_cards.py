"""Build hashed, bounded source cards from an immutable local corpus."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import shutil
import sqlite3
import subprocess
from pathlib import Path
from typing import Any, Sequence

from fitz_tool.contracts import validate_source_card


EXTENSION_MODALITIES = {
    ".md": "text",
    ".txt": "text",
    ".rst": "text",
    ".yaml": "text",
    ".yml": "text",
    ".json": "text",
    ".csv": "csv",
    ".xlsx": "excel",
    ".xlsm": "excel",
    ".sqlite": "sqlite",
    ".sqlite3": "sqlite",
    ".db": "sqlite",
    ".pdf": "pdf",
    ".py": "code",
    ".js": "code",
    ".ts": "code",
}


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--root", action="append", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--max-facts", type=int, default=80)
    parser.add_argument("--ignore-unsupported", action="store_true")
    return parser


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _normalized_sha256(path: Path, modality: str) -> str | None:
    if modality not in {"text", "code"}:
        return None
    content = path.read_text(encoding="utf-8", errors="replace").strip()
    return hashlib.sha256(content.encode("utf-8")).hexdigest()


def _text_blocks(content: str, max_facts: int) -> list[tuple[str, str]]:
    blocks = re.split(r"\n\s*\n", content)
    facts: list[tuple[str, str]] = []
    for index, block in enumerate(blocks, start=1):
        statement = " ".join(line.strip() for line in block.splitlines() if line.strip())
        if statement:
            facts.append((statement, f"block:{index}"))
        if len(facts) == max_facts:
            break
    return facts


def _pdf_facts(path: Path, max_facts: int) -> tuple[list[tuple[str, str]], str]:
    try:
        from pypdf import PdfReader  # type: ignore[import-not-found]

        pages = PdfReader(str(path)).pages
        facts: list[tuple[str, str]] = []
        for page_number, page in enumerate(pages, start=1):
            facts.extend(
                (statement, f"page:{page_number}:block:{index}")
                for index, (statement, _locator) in enumerate(
                    _text_blocks(page.extract_text() or "", max_facts), start=1
                )
            )
            if len(facts) >= max_facts:
                break
        return facts[:max_facts], "pypdf"
    except ImportError:
        executable = shutil.which("pdftotext")
        if not executable:
            raise RuntimeError("PDF extraction requires pypdf or pdftotext")
        completed = subprocess.run(
            [executable, "-layout", str(path), "-"],
            capture_output=True,
            text=True,
            check=False,
        )
        if completed.returncode:
            raise RuntimeError(completed.stderr.strip() or "pdftotext failed")
        pages = completed.stdout.split("\f")
        facts = []
        for page_number, page in enumerate(pages, start=1):
            facts.extend(
                (statement, f"page:{page_number}:block:{index}")
                for index, (statement, _locator) in enumerate(
                    _text_blocks(page, max_facts), start=1
                )
            )
            if len(facts) >= max_facts:
                break
        return facts[:max_facts], "pdftotext"


def _csv_facts(path: Path, max_facts: int) -> tuple[list[tuple[str, str]], str]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.reader(handle))
    if not rows:
        return [], "csv"
    headers = rows[0]
    facts = [(f"Columns: {', '.join(headers)}", "header")]
    for row_number, row in enumerate(rows[1:], start=2):
        values = ", ".join(
            f"{header}={value}" for header, value in zip(headers, row, strict=False) if value
        )
        if values:
            facts.append((f"Row {row_number}: {values}", f"row:{row_number}"))
        if len(facts) == max_facts:
            break
    return facts, "csv"


def _sqlite_facts(path: Path, max_facts: int) -> tuple[list[tuple[str, str]], str]:
    uri = f"file:{path.as_posix()}?mode=ro"
    connection = sqlite3.connect(uri, uri=True)
    try:
        tables = [row[0] for row in connection.execute("SELECT name FROM sqlite_master WHERE type='table'")]
        facts: list[tuple[str, str]] = []
        for table in tables:
            columns = connection.execute(f'PRAGMA table_info("{table}")').fetchall()
            names = ", ".join(str(row[1]) for row in columns)
            facts.append((f"Table {table} has columns: {names}", f"table:{table}:schema"))
            rows = connection.execute(f'SELECT * FROM "{table}" LIMIT 10').fetchall()
            for row_number, row in enumerate(rows, start=1):
                facts.append((f"Table {table} sample row {row_number}: {row}", f"table:{table}:row:{row_number}"))
                if len(facts) == max_facts:
                    return facts, "sqlite-readonly"
        return facts, "sqlite-readonly"
    finally:
        connection.close()


def _xlsx_facts(path: Path, max_facts: int) -> tuple[list[tuple[str, str]], str]:
    try:
        from openpyxl import load_workbook  # type: ignore[import-not-found]
    except ImportError as exc:
        raise RuntimeError("Excel extraction requires openpyxl") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    facts: list[tuple[str, str]] = []
    try:
        for sheet in workbook.worksheets:
            rows = sheet.iter_rows(values_only=True)
            header = next(rows, None)
            if header:
                names = ", ".join(str(value) for value in header if value is not None)
                facts.append((f"Sheet {sheet.title} columns: {names}", f"sheet:{sheet.title}:header"))
            for row_number, row in enumerate(rows, start=2):
                values = ", ".join(str(value) for value in row if value is not None)
                if values:
                    facts.append((f"Sheet {sheet.title} row {row_number}: {values}", f"sheet:{sheet.title}:row:{row_number}"))
                if len(facts) == max_facts:
                    return facts, "openpyxl-readonly"
        return facts, "openpyxl-readonly"
    finally:
        workbook.close()


def _facts(path: Path, modality: str, max_facts: int) -> tuple[list[tuple[str, str]], str]:
    if modality == "text" or modality == "code":
        return _text_blocks(path.read_text(encoding="utf-8", errors="replace"), max_facts), "text-lines"
    if modality == "pdf":
        return _pdf_facts(path, max_facts)
    if modality == "csv":
        return _csv_facts(path, max_facts)
    if modality == "sqlite":
        return _sqlite_facts(path, max_facts)
    if modality == "excel":
        return _xlsx_facts(path, max_facts)
    raise ValueError(f"unsupported modality {modality!r}")


def _card(root: Path, path: Path, max_facts: int) -> dict[str, Any]:
    content_hash = _sha256(path)
    modality = EXTENSION_MODALITIES[path.suffix.casefold()]
    facts, extraction_method = _facts(path, modality, max_facts)
    if not facts:
        raise ValueError(f"no extractable facts in {path}")
    relative = path.relative_to(root).as_posix()
    card = {
        "schema_version": "source-card.v1",
        "source_id": relative,
        "document_id": "doc-" + content_hash[:16],
        "title": path.stem.replace("_", " ").replace("-", " ").title(),
        "modality": modality,
        "content_sha256": content_hash,
        "source_path": relative,
        "extraction_method": extraction_method,
        "facts": [
            {
                "fact_id": f"{content_hash[:12]}-f{index:04d}",
                "statement": statement,
                "locator": locator,
            }
            for index, (statement, locator) in enumerate(facts, start=1)
        ],
        "metadata": {"root": str(root), "immutable_input": True},
    }
    normalized_hash = _normalized_sha256(path, modality)
    if normalized_hash:
        card["normalized_content_sha256"] = normalized_hash
    return card


def main(argv: Sequence[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    if args.max_facts < 1:
        raise SystemExit("max-facts must be positive")
    cards: list[dict[str, Any]] = []
    errors: list[dict[str, str]] = []
    for root in args.root:
        root = root.resolve()
        if not root.is_dir():
            raise SystemExit(f"source root is not a directory: {root}")
        for path in sorted(item for item in root.rglob("*") if item.is_file()):
            modality = EXTENSION_MODALITIES.get(path.suffix.casefold())
            if modality is None:
                if args.ignore_unsupported:
                    continue
                errors.append({"path": str(path), "error": "unsupported file extension"})
                continue
            try:
                card = _card(root, path, args.max_facts)
                report = validate_source_card(card)
                if report.valid:
                    cards.append(card)
                else:
                    errors.append({"path": str(path), "error": json.dumps(report.as_dict(), sort_keys=True)})
            except (OSError, RuntimeError, ValueError, sqlite3.Error) as exc:
                errors.append({"path": str(path), "error": str(exc)})

    args.output.parent.mkdir(parents=True, exist_ok=True)
    with args.output.open("w", encoding="utf-8", newline="\n") as handle:
        for card in cards:
            handle.write(json.dumps(card, ensure_ascii=False, sort_keys=True) + "\n")
    if errors:
        args.output.with_suffix(".errors.jsonl").write_text(
            "".join(json.dumps(error, sort_keys=True) + "\n" for error in errors),
            encoding="utf-8",
        )
    print(json.dumps({"cards": len(cards), "errors": len(errors), "output": str(args.output)}, indent=2, sort_keys=True))
    return 0 if not errors else 1


if __name__ == "__main__":
    raise SystemExit(main())
